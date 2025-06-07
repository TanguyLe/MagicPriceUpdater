from typing import List, Optional, Tuple, Union

import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

EXCEL_ENGINE = "openpyxl"


def get_width_value(centimeter_value: float):
    return centimeter_value / 0.24706791635539999


def is_okay_col(
    col_name: Union[str, Tuple[Optional[str], Optional[str], Optional[str]]],
    df_index_name: Tuple[str, str, str],
) -> bool:
    if isinstance(col_name, str):
        if isinstance(df_index_name, str):
            return col_name == df_index_name

        return col_name in df_index_name

    if isinstance(df_index_name, str):
        raise ValueError("can't use multiIndex col name with df without multi index")

    return all(
        [
            col_n is None or col_n == df_index_n
            for col_n, df_index_n in zip(col_name, df_index_name)
        ]
    )


def get_excel_col_names(df: pd.DataFrame, col_name: str) -> List[str]:
    """Returns the letter to access the excel position of a column"""
    return [
        get_column_letter(index + 2)
        for index, col in enumerate(list(df.columns))
        if is_okay_col(col_name=col_name, df_index_name=col)
    ]


def change_visibility(worksheet, excel_col_name: str, property_value):
    worksheet.column_dimensions[excel_col_name].hidden = property_value


def color_fill(worksheet, excel_col_name: str, property_value):
    fill = PatternFill(
        start_color=property_value, end_color=property_value, fill_type="solid"
    )

    worksheet.column_dimensions[excel_col_name].fill = fill


def change_width(worksheet, excel_col_name: str, property_value):
    worksheet.column_dimensions[excel_col_name].width = get_width_value(
        centimeter_value=property_value
    )


PROPERTIES_MAP = {
    "hidden": change_visibility,
    "color": color_fill,
    "width": change_width,
}


def format_excel_df(
    df: pd.DataFrame, writer: pd.ExcelWriter, format_config: dict, sheet_name="Sheet1"
):
    worksheet = writer.sheets[sheet_name]

    for col_name, format_properties in format_config.items():
        excel_col_names = get_excel_col_names(df=df, col_name=col_name)
        for excel_col_name in excel_col_names:
            for property_name, property_value in format_properties.items():
                PROPERTIES_MAP[property_name](
                    worksheet=worksheet,
                    excel_col_name=excel_col_name,
                    property_value=property_value,
                )
