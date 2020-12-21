import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


def get_width_value(centimeter_value: float):
    return centimeter_value / 0.24706791635539999


def get_excel_col_name(df: pd.DataFrame, col_name: str) -> str:
    """Returns the letter to access the excel position of a column"""
    return get_column_letter(list(df.columns).index(col_name) + 1)


def change_visibility(worksheet, excel_col_name: str, property_value):
    worksheet.column_dimensions[excel_col_name].hidden = property_value


def color_fill(worksheet, excel_col_name: str, property_value):
    fill = PatternFill(start_color=property_value, end_color=property_value, fill_type='solid')

    worksheet.column_dimensions[excel_col_name].fill = fill


def change_width(worksheet, excel_col_name: str, property_value):
    worksheet.column_dimensions[excel_col_name].width = get_width_value(
        centimeter_value=property_value
    )


PROPERTIES_MAP = {
    "hidden": change_visibility,
    "color": color_fill,
    "width": change_width
}


def format_and_save_df(df: pd.DataFrame, writer: pd.ExcelWriter, format_config: dict):
    worksheet = writer.sheets['Sheet1']

    for col_name, format_properties in format_config.items():
        excel_col_name = get_excel_col_name(df=df, col_name=col_name)
        for property_name, property_value in format_properties.items():
            PROPERTIES_MAP[property_name](
                worksheet=worksheet, excel_col_name=excel_col_name, property_value=property_value
            )
    writer.save()
