import base64
import gzip
import io
from pathlib import Path

import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

HIDDEN_COLS = ["idArticle", "Local Name", "Exp. Name", "Signed?", "Playset?", "Altered?", "idCurrency", "Currency Code"]
COLORED_COLS = {
    "Price": "949494E8",
    "SuggestedPrice": "949494E8",
    "PriceApproval": "FFF0F8FF"
}


def get_excel_col_name(df: pd.DataFrame, col_name: str) -> str:
    """Returns the letter to access the excel position of a column"""
    return get_column_letter(list(df.columns).index(col_name) + 1)


def get_stock_file_path(folder_path: Path) -> Path:
    """Constructs the stock file path from a folder path"""
    return folder_path / "stock.xlsx"


def convert_base64_gzipped_string_to_dataframe(b64_zipped_string: str) -> pd.DataFrame:
    """Converts a base64 gzipped strings (or bytes) into a pandas dataframe"""
    decoded_string = base64.b64decode(b64_zipped_string)
    csv_string = gzip.decompress(decoded_string)

    return pd.read_csv(io.BytesIO(csv_string), sep=";")


def save_stock_df_as_odf_formatted_file(df: pd.DataFrame, file_path: Path) -> None:
    writer = pd.ExcelWriter(path=str(file_path))
    df.to_excel(excel_writer=writer, index=False)

    worksheet = writer.sheets['Sheet1']

    for col_name in HIDDEN_COLS:
        excel_col_name = get_excel_col_name(df=df, col_name=col_name)
        worksheet.column_dimensions[excel_col_name].hidden = True

    for col_name, color in COLORED_COLS.items():
        excel_col_name = get_excel_col_name(df=df, col_name=col_name)
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

        worksheet.column_dimensions[excel_col_name].fill = fill

    excel_col_name = get_excel_col_name(df=df, col_name="English Name")
    worksheet.column_dimensions[excel_col_name].width = 50

    writer.save()
